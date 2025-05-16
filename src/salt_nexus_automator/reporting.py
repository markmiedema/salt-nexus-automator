# src/reporting.py
import pandas as pd
import os
import json
import logging
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as ReportlabImage, PageBreak, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.units import inch
from reportlab.lib import colors
from .utils import ErrorCollector

class ReportGenerator:
    """Generates client-ready Excel and PDF reports, plus error/summary files."""

    def __init__(self, client_name: str, error_collector: ErrorCollector,
                 output_dir: str = "output", logo_path: str = None):
        """
        Initializes the ReportGenerator.

        Args:
            client_name (str): Name of the client for branding.
            error_collector (ErrorCollector): Instance containing run results.
            output_dir (str): Directory to save reports. Defaults to "output".
            logo_path (str, optional): Path to the company logo file. Defaults to None.
        """
        self.client_name = client_name
        self.error_collector = error_collector
        self.output_dir = output_dir
        self.logo_path = logo_path
        self.timestamp = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S') # Include time for uniqueness
        # Use timestamp from error_collector if available for consistency
        if hasattr(error_collector, 'summary') and 'start_time' in error_collector.summary:
             self.timestamp = error_collector.summary['start_time'].strftime('%Y%m%d_%H%M%S')

        self.base_filename = f"{self.client_name.replace(' ', '_')}_Nexus_Analysis_{self.timestamp}"
        os.makedirs(self.output_dir, exist_ok=True)
        logging.info(f"ReportGenerator initialized. Output directory: {self.output_dir}")

    def _apply_excel_styles(self, ws):
        """Applies basic formatting to an Excel worksheet."""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))

        # Style header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        # Auto-adjust column widths and apply borders/alignment to data rows
        for col_idx, column_cells in enumerate(ws.columns):
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            # Calculate max width needed
            for cell in column_cells:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                # Apply border and alignment to data cells (skip header)
                if cell.row > 1:
                    cell.border = thin_border
                    # Align numeric columns to the right? Check dtype if needed.
                    # For simplicity, left-align all data for now.
                    cell.alignment = left_alignment

            adjusted_width = (max_length + 2) * 1.2 # Add padding
            ws.column_dimensions[column_letter].width = min(adjusted_width, 60) # Limit max width


    def generate_excel_report(self, data_dict: dict):
        """
        Generates a multi-sheet Excel report with improved formatting.

        Args:
            data_dict (dict): A dictionary where keys are sheet names and values are DataFrames.
        """
        excel_path = os.path.join(self.output_dir, f"{self.base_filename}.xlsx")
        logging.info(f"Generating Excel report: {excel_path}")

        try:
            with pd.ExcelWriter(excel_path, engine='openpyxl',
                                datetime_format='YYYY-MM-DD',
                                date_format='YYYY-MM-DD') as writer:
                wb = writer.book
                first_sheet = True

                # --- Cover Sheet ---
                ws_cover = wb.create_sheet("Cover", 0) # Insert as first sheet
                ws_cover['B2'] = "SALT Economic Nexus Analysis"
                ws_cover['B2'].font = Font(size=16, bold=True)
                ws_cover['B4'] = f"Client:"
                ws_cover['C4'] = self.client_name
                ws_cover['B5'] = f"Report Date:"
                ws_cover['C5'] = pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')
                # Add Summary Stats
                summary = self.error_collector.get_summary()
                ws_cover['B7'] = "Run Summary:"
                ws_cover['B7'].font = Font(bold=True)
                row = 8
                for key, value in summary.items():
                    if key not in ['start_time', 'end_time', 'warnings']: # Exclude raw lists/timestamps
                        ws_cover[f'B{row}'] = f"{key.replace('_', ' ').title()}:"
                        ws_cover[f'C{row}'] = value if not isinstance(value, list) else len(value)
                        row += 1
                # Add Logo
                if self.logo_path and os.path.exists(self.logo_path):
                     try:
                         img = OpenpyxlImage(self.logo_path)
                         img.height = 75 # Adjust size as needed
                         img.width = 150
                         ws_cover.add_image(img, 'E2')
                     except Exception as e:
                         logging.warning(f"Could not embed logo in Excel: {e}")
                # Set column widths for cover sheet
                ws_cover.column_dimensions['B'].width = 30
                ws_cover.column_dimensions['C'].width = 30

                # --- Data Sheets ---
                for sheet_name, df_orig in data_dict.items():
                    if isinstance(df_orig, pd.DataFrame) and not df_orig.empty:
                        logging.debug(f"Writing sheet: {sheet_name}")
                        # Convert Period columns to string for Excel compatibility
                        df = df_orig.copy()
                        for col in df.select_dtypes(include=['period']).columns:
                             df[col] = df[col].astype(str)

                        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
                        ws = writer.sheets[sheet_name]
                        self._apply_excel_styles(ws)
                        # Freeze top row
                        ws.freeze_panes = 'A2'

                    elif isinstance(df_orig, pd.DataFrame) and df_orig.empty:
                         logging.info(f"Skipping empty DataFrame for sheet: {sheet_name}")
                         # Create an empty sheet with a note
                         ws = wb.create_sheet(sheet_name)
                         ws['A1'] = "No data available for this section."
                    else:
                         logging.warning(f"Item for sheet '{sheet_name}' is not a DataFrame or is invalid. Skipping.")

            logging.info("Excel report generated successfully.")

        except Exception as e:
             logging.error(f"Failed to generate Excel report: {e}", exc_info=True)


    def generate_pdf_report(self, summary_text: str = "Refer to the accompanying Excel file for detailed results."):
        """Generates a simple PDF cover page and potentially a summary table."""
        pdf_path = os.path.join(self.output_dir, f"{self.base_filename}.pdf")
        logging.info(f"Generating PDF report: {pdf_path}")

        try:
            doc = SimpleDocTemplate(pdf_path, pagesize=(inch*8.5, inch*11),
                                    leftMargin=0.75*inch, rightMargin=0.75*inch,
                                    topMargin=0.75*inch, bottomMargin=0.75*inch)
            styles = getSampleStyleSheet()
            story = []

            # --- PDF Cover Page Elements ---
            # Logo
            if self.logo_path and os.path.exists(self.logo_path):
                try:
                    img = ReportlabImage(self.logo_path, width=2*inch, height=1*inch)
                    img.hAlign = 'LEFT'
                    story.append(img)
                    story.append(Spacer(1, 0.3*inch))
                except Exception as e:
                    logging.warning(f"Could not embed logo in PDF: {e}")

            # Title and Client Info
            story.append(Paragraph("SALT Economic Nexus Analysis", styles['h1']))
            story.append(Spacer(1, 0.2*inch))
            story.append(Paragraph(f"Client: {self.client_name}", styles['h2']))
            story.append(Paragraph(f"Report Date: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['normal']))
            story.append(Spacer(1, 0.5*inch))
            story.append(Paragraph(summary_text, styles['normal']))
            story.append(Spacer(1, 0.3*inch))

            # --- PDF Summary Table (Example: Key Stats) ---
            summary = self.error_collector.get_summary()
            summary_data_for_pdf = [
                ["Parameter", "Value"],
                ["Total Rows Input", summary.get('total_rows_input', 'N/A')],
                ["Rows Processed", summary.get('rows_processed', 'N/A')],
                ["Rows Rejected", summary.get('rows_rejected', 'N/A')],
                ["Nexus Triggers (States)", summary.get('nexus_triggers', 'N/A')],
                ["Exposure Calculated (States)", summary.get('states_with_exposure', 'N/A')],
                ["Warnings Issued", summary.get('warnings_count', 'N/A')],
                ["Run Duration (Seconds)", f"{summary.get('duration_seconds', 0):.2f}"]
            ]

            summary_table = Table(summary_data_for_pdf, colWidths=[3*inch, 3*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.grey),
                ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0,0), (-1,0), 12),
                ('BACKGROUND', (0,1), (-1,-1), colors.beige),
                ('GRID', (0,0), (-1,-1), 1, colors.black)
            ]))
            story.append(Paragraph("Run Summary:", styles['h3']))
            story.append(Spacer(1, 0.1*inch))
            story.append(summary_table)

            # TODO: Optionally add more summary tables (e.g., top triggered states) or charts to story[]

            # --- Build PDF ---
            doc.build(story)
            logging.info("PDF report generated successfully.")

        except Exception as e:
            logging.error(f"Error generating PDF report: {e}", exc_info=True)


    def generate_error_reports(self):
        """Generates the rejected rows CSV and the run summary JSON."""
        logging.info("Generating error reports (Rejected Rows CSV and Run Summary JSON)...")
        # Rejected Rows CSV
        reject_df = self.error_collector.get_rejected_rows_df()
        reject_filename = f"rejected_rows_{self.timestamp}.csv"
        reject_path = os.path.join(self.output_dir, reject_filename)
        if not reject_df.empty:
            try:
                reject_df.to_csv(reject_path, index=False, encoding='utf-8')
                logging.info(f"Rejected rows report saved to: {reject_path} ({len(reject_df)} rows)")
            except Exception as e:
                logging.error(f"Failed to save rejected rows report: {e}", exc_info=True)
        else:
            logging.info("No rejected rows to report.")
            # Optionally create an empty file or just skip
            # with open(reject_path, 'w') as f: f.write("No rejected rows.\n")

        # Run Summary JSON (Finalize summary before getting)
        self.error_collector.finalize_summary() # Calculate duration etc.
        summary_data = self.error_collector.get_summary()
        summary_filename = f"run_summary_{self.timestamp}.json"
        summary_path = os.path.join(self.output_dir, summary_filename)
        try:
            with open(summary_path, 'w', encoding='utf-8') as f:
                # Use default=str for any non-standard JSON types (though finalize converts times)
                json.dump(summary_data, f, indent=4, default=str)
            logging.info(f"Run summary report saved to: {summary_path}")
        except Exception as e:
            logging.error(f"Failed to save run summary report: {e}", exc_info=True)


    def generate_all_reports(self, data_dict: dict, pdf_summary: str = None):
        """Generates all standard reports: Excel, PDF (optional), Error/Summary."""
        logging.info("Starting generation of all reports...")
        # Generate main data report first
        self.generate_excel_report(data_dict)

        # Generate optional PDF
        if pdf_summary:
            self.generate_pdf_report(summary_text=pdf_summary)
        else:
             # Generate a default simple PDF cover even if no specific text given
             self.generate_pdf_report()


        # Always generate error/summary files last, after finalize_summary is called
        self.generate_error_reports()
        logging.info("All report generation tasks complete.")